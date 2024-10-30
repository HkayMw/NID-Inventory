from kivy.uix.screenmanager import Screen
# from controller.user_controller import UserController
from kivy.clock import Clock
from kivy.lang import Builder
from kivy.core.window import Window
from kivy.uix.boxlayout import BoxLayout
from kivymd.app import MDApp
from kivy.metrics import dp
from kivymd.uix.datatables import MDDataTable
from kivy.uix.anchorlayout import AnchorLayout
from datetime import datetime
from openpyxl import Workbook
import os
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from kivymd.uix.pickers import MDDatePicker

from controller.report_controller import ReportController

# Builder.load_file('view/UI.kv')
Builder.load_file('view/admin/report/report_view.kv')


class ReportScreen(Screen):
    def __init__(self, **kwargs):
        super(ReportScreen, self).__init__(**kwargs)
        # self.controller = UserController()
        self.report_controller = ReportController()
        self.data_tables = None
        self.app = MDApp.get_running_app()
        self.notice = self.app.root.ids.notice
        
        Clock.schedule_once(self.initialize_table, 0.2)
        # Bind text changes of start_date and end_date fields to load_report
        self.ids.start_date.bind(text=self.load_report)
        self.ids.end_date.bind(text=self.load_report)

    def on_enter(self, *args):
        # print(self.parent.current)
        self.current_user = self.app.user_details
        self.user_id = self.current_user['id_number']
        
        # Set today's date in the 'start_date' and 'end_date' fields
        today_date = datetime.now().strftime('%Y-%m-%d')
        self.ids.start_date.text = today_date
        self.ids.end_date.text = today_date
        
    def initialize_table(self, *args):
        # Initialize the table with default headers and an empty state
        layout = AnchorLayout()
        self.data_tables = MDDataTable(
            
            # use_pagination=True,
            rows_num = 8,
            # check=True,
            # name column, width column, sorting function column(optional), custom tooltip
            column_data = [
                    ("Metric", dp(70)),
                    ("Value", dp(30)),
                    # ("Average", dp(30)),
                    # ("Weekly Average", dp(30)),
                    # ("Monthly Average", dp(30)),
                ],
            row_data = [
                    ["IDs in Storage", "-"],
                    ["IDs Added", "-"],
                    ["IDs Collected", "-"],
                    ["Notifications Sent", "-"],
                    ["Number of IDs Collected After Notification", "-"],
                    # ["Notification Success Rate", "-"],
                    ["Collection % for Notified Clients", "-"],
                    ["Collection % for Non-Notified Clients", "-"],
                ]
        )
       
        # self.ids.lastname_field.text = "something"
        self.ids.table_container.clear_widgets()
        layout.add_widget(self.data_tables)
        self.ids.table_container.add_widget(layout)
        
        # self.data_tables.bind(on_row_press=self.on_row_press)
        
    # def calculate_metrics(self, report_data, start_date, end_date):
    #     # Calculate the number of days between start_date and end_date
    #     start = datetime.strptime(start_date, "%Y-%m-%d")
    #     end = datetime.strptime(end_date, "%Y-%m-%d")
    #     total_days = (end - start).days + 1  # Include both start and end dates
        
    #     # Set total_days to 1 if start_date == end_date
    #     total_days = max(total_days, 1)  # Ensure at least 1 day for averaging
        
    #     metrics = {
    #         'average_ids_added_per_day': report_data['ids_added'] / total_days,
    #         'average_ids_collected_per_day': report_data['ids_issued'] / total_days,
    #         # 'average_ids_in_storage_per_day': report_data['ids_in_storage'] / total_days,
    #         'average_notifications_sent_per_day': report_data['sent_sms'] / total_days,
    #         'average_notified_ids_collected_per_day': report_data['notified_ids_issued'] / total_days,
    #         'collection_percentage_notified': (report_data['notified_ids_issued'] / report_data['ids_issued'] * 100) if report_data['ids_issued'] > 0 else 0,
    #         'collection_percentage_non_notified': ((report_data['ids_issued'] - report_data['notified_ids_issued']) / report_data['ids_issued'] * 100) if report_data['ids_issued'] > 0 else 0,
    #     }

    #     return metrics
    
    def populate_table(self, report_data, start_date, end_date):
        # Calculate the metrics
        # metrics = self.calculate_metrics(report_data, start_date, end_date)
        # start = datetime.strptime(start_date, "%Y-%m-%d")
        # end = datetime.strptime(end_date, "%Y-%m-%d")
        # total_days = (end - start).days + 1  # Include both start and end dates

        # Create a new list of lists for row data
        new_row_data = [
            ["IDs in Storage", report_data['ids_in_storage']],
            ["IDs Added", report_data['ids_added']],
            ["IDs Collected", report_data['ids_issued']],
            ["Notifications Sent", report_data['sent_sms']],
            ["Number of IDs Collected After Notification",  report_data['notified_ids_issued']],
            ["Collection % for Notified Clients", f'{(report_data['notified_ids_issued'] / report_data['ids_issued'] * 100) if report_data['ids_issued'] > 0 else 0:.2f}%'],
            ["Collection % for Non-Notified Clients", f'{((report_data['ids_issued'] - report_data['notified_ids_issued']) / report_data['ids_issued'] * 100) if report_data['ids_issued'] > 0 else 0:.2f}%'],
        ]

        # Set the new row data to the MDDataTable
        self.data_tables.row_data = new_row_data
       
        
    def show_date_picker(self, text_field):
        # Create the date picker
        date_dialog = MDDatePicker(
            primary_color=self.app.theme_cls.primary_color,
            selector_color=self.app.theme_cls.accent_color,
            text_button_color=self.app.theme_cls.primary_dark,
        )

        # Set the selected date callback with text_field
        date_dialog.bind(on_save=lambda instance, value, date_range: self.on_save_date(text_field, value))

        # Open the date picker
        date_dialog.open()

    def on_save_date(self, text_field, value):
        # Format the selected date
        formatted_date = value.strftime('%Y-%m-%d')
        text_field.text = formatted_date

        start_date_str = self.ids.start_date.text
        end_date_str = self.ids.end_date.text
        if text_field.name == 'end_date':
            if (start_date_str and end_date_str):
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d')  # Parse start date
                end_date = datetime.strptime(formatted_date, '%Y-%m-%d')  # Parse end date
                
                
                
                # Check if end date is less than start date
                if end_date < start_date:
                    # Set start date to end date 
                    print('end is less')
                    self.ids.start_date.text = formatted_date
                    self.ids.end_date.text = formatted_date      
                
        # If the text field is 'start_date', check if it is greater than 'end_date'
        if text_field.name == 'start_date':
            # start_date_str = self.ids.start_date.text
            if (start_date_str and end_date_str):
                start_date = datetime.strptime(formatted_date, '%Y-%m-%d')  # Parse start date
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d')  # Parse end date
                
                # Check if end date is less than start date
                if  start_date > end_date:
                    # Set start date to end date
                    self.ids.end_date.text = formatted_date      
                    self.ids.start_date.text = formatted_date
     
    def load_report(self, *args):
        start_date = self.ids.start_date.text
        end_date = self.ids.end_date.text
        
        if not start_date or not end_date:
            return
        # start = datetime.strptime(start_date, "%Y-%m-%d")
        # end = datetime.strptime(end_date, "%Y-%m-%d")
        
        # print(start_date)
        # print(end_date)
        # print(start)
        # print(end)
        
        # success, message, data = self.report_controller.load_report(start, end)
        success, message, data = self.report_controller.load_report(start_date, end_date)
        if success:
            # print(data)
            self.populate_table(data, start_date, end_date)
        else:
            print(message)
        
         
    def export_report(self, start_date, end_date):
        # Get the path to the user's Documents folder
        documents_dir = os.path.expanduser("~/Documents")
        reports_dir = os.path.join(documents_dir, 'Reports')

        # Ensure the 'Reports' directory exists
        os.makedirs(reports_dir, exist_ok=True)

        # Sanitize the date format for the filename
        sanitized_start_date = start_date.replace("/", "-")
        sanitized_end_date = end_date.replace("/", "-")
        base_file_name = f"mzuzu_report_{sanitized_start_date}_{sanitized_end_date}"
        file_name = f"{base_file_name}.xlsx"
        file_path = os.path.join(reports_dir, file_name)

        # Check if the file already exists and modify the filename accordingly
        counter = 1
        while os.path.exists(file_path):
            file_name = f"{base_file_name} ({counter}).xlsx"
            file_path = os.path.join(reports_dir, file_name)
            counter += 1

        # Reference your data table instance (adjust to match your data table ID)
        data_table = self.data_tables

        # Create a new workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Report"

        # Define a border style
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        # Add title and subtitles
        title = "MZUZU NRB"
        subtitle1 = "ID Inventory Manager Report"
        subtitle2 = f"Dates: {start_date} to {end_date}"

        # Write the title
        sheet.merge_cells('A1:B1')  # Adjust the range as needed
        sheet['A1'] = title
        sheet['A1'].font = Font(bold=True, size=16)  # Title font settings
        sheet['A1'].alignment = Alignment(horizontal='center')  # Center align title
        # sheet['A1'].border = thin_border  # Add border to title cell

        # Write the first subtitle
        sheet.merge_cells('A2:B2')  # Adjust the range as needed
        sheet['A2'] = subtitle1
        sheet['A2'].font = Font(bold=True, size=14)  # Subtitle font settings
        sheet['A2'].alignment = Alignment(horizontal='center')  # Center align subtitle
        # sheet['A2'].border = thin_border  # Add border to first subtitle cell

        # Write the second subtitle
        sheet.merge_cells('A3:B3')  # Adjust the range as needed
        sheet['A3'] = subtitle2
        sheet['A3'].font = Font(bold=True, size=12)  # Subtitle font settings
        sheet['A3'].alignment = Alignment(horizontal='center')  # Center align subtitle
        # sheet['A3'].border = thin_border  # Add border to second subtitle cell

        # Write the header row using column names from the data table
        for col_index, column_header in enumerate(data_table.column_data, start=1):
            cell = sheet.cell(row=4, column=col_index, value=column_header[0])
            cell.font = Font(bold=True)  # Make headers bold
            cell.border = thin_border  # Add border to header cells

        # Write the data rows from the data table and adjust column width
        for col_index in range(1, len(data_table.column_data) + 1):
            max_length = len(data_table.column_data[col_index - 1][0])  # Header length
            for row_index, row_data in enumerate(data_table.row_data, start=5):  # Start from row 4 for data
                cell_value = row_data[col_index - 1]  # Adjust for zero-based index
                cell_length = len(str(cell_value))  # Calculate length of the cell value
                max_length = max(max_length, cell_length)  # Find max length

                cell = sheet.cell(row=row_index, column=col_index, value=cell_value)
                cell.border = thin_border  # Add border to data cells

            # Set the column width based on max length
            adjusted_width = max_length + 2  # Add a little extra space
            sheet.column_dimensions[get_column_letter(col_index)].width = adjusted_width

        # Save the workbook to the 'Reports' directory in the Documents folder
        workbook.save(file_path)
        self.notice.text =f"Report saved as {file_name}"
